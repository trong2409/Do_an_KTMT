import psutil
import openpyxl
import time

pid = input("Input PID: ")

excel_file_name = f'Process_info_{time.strftime("%H_%M_%S")}.xlsx'

wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "Time"
ws["B1"] = "CPU usage (%)"
ws["C1"] = "Totalmemory usage (%))"

process = psutil.Process(int(pid))

i = 2
while True:
    cpu_usage = process.cpu_percent()
    ram_usage = process.memory_percent()

    ws["A" + str(i)] = time.strftime("%H:%M:%S")
    ws["B" + str(i)] = cpu_usage
    ws["C" + str(i)] = ram_usage
    print(cpu_usage, "--", ram_usage)

    wb.save(excel_file_name)

    time.sleep(1)

    i += 1
